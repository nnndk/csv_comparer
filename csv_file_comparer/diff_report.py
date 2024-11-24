from enum import Enum
from openpyxl import *
from openpyxl.styles import Color, PatternFill, Font


class DiffReport:
    def __init__(self):
        self._report_rows: list[DiffReportRow] = []
        self._columns_width: dict[int, int] = {}


    def append_header_row(self, row: list[str]):
        row = DiffReportRow(RowType.HEADER, row)
        self._report_rows.append(row)


    def append_added_row(self, row: list[str]):
        row = DiffReportRow(RowType.ADDED, row)
        self._report_rows.append(row)


    def append_modified_row(self, row: list[str], modified_column_ids: list[int]):
        row = DiffReportRow(RowType.MODIFIED, row, modified_column_ids)
        self._report_rows.append(row)


    def append_removed_row(self, row: list[str]):
        row = DiffReportRow(RowType.REMOVED, row)
        self._report_rows.append(row)


    def create_diff_report(self, report_path: str):
        wb = Workbook()
        ws = wb.active
        current_row = 0

        for report_row in self._report_rows:
            ws.append(report_row.row)
            current_row += 1

            if report_row.row_type == RowType.MODIFIED:
                for column in report_row.modified_column_ids:
                    color_by_row_type = report_row.get_row_color()
                    color_fill = PatternFill(patternType='solid', fgColor=color_by_row_type)
                    ws.cell(current_row, column).fill = color_fill
            elif (report_row.row_type == RowType.ADDED) or (report_row.row_type == RowType.REMOVED):
                for column in range(1, len(report_row.row) + 1):
                    color_by_row_type = report_row.get_row_color()
                    color_fill = PatternFill(patternType='solid', fgColor=color_by_row_type)
                    ws.cell(current_row, column).fill = color_fill
            elif report_row.row_type == RowType.HEADER:
                for column in range(1, len(report_row.row) + 1):
                    ws.cell(current_row, column).font = Font(bold=True)

            for i in range(1, len(report_row.row) + 1):
                if i in self._columns_width.keys():
                    self._columns_width[i] = max(self._columns_width[i], len(report_row.row[i - 1]))
                else:
                    self._columns_width[i] = len(report_row.row[i - 1])

        for k, v in self._columns_width.items():
            column_letter = ws.cell(1, k).column_letter
            ws.column_dimensions[column_letter].width = v + 3

        wb.save(report_path)


class RowType(Enum):
    HEADER = 0
    ADDED = 1
    MODIFIED = 2
    REMOVED = 3


class DiffReportRow:
    ROW_TYPE_COLORS = {
        RowType.ADDED: Color("b6d7a8"),
        RowType.MODIFIED: Color("ffff00"),
        RowType.REMOVED: Color("ed8886"),
    }

    def __init__(self, row_type: RowType, row: list[str], modified_column_ids: list[int] = None):
        self.row_type: RowType = row_type
        self.row: list[str] = row
        self.modified_column_ids: list[int] = modified_column_ids # only for RowType.MODIFIED


    def get_row_color(self):
        return DiffReportRow.ROW_TYPE_COLORS[self.row_type]
